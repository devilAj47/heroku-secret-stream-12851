from flask import Flask, render_template, Response, jsonify, url_for, make_response
import json
from camera import VideoCamera

app = Flask(__name__)
import io
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.backends.backend_svg import FigureCanvasSVG
import numpy as np
import cv2 as cv
from keras.models import load_model
from operator import add
# from tkinter import Tk, mainloop, TOP
# from tkinter import messagebox

# from tkinter import Tk, mainloop, TOP 
# from tkinter.ttk import Button 
from time import time
# from tkinter import * 
import matplotlib.pyplot as plt
# from tkinter import *
# from tkinter.ttk import *

import matplotlib

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import xlsxwriter

from matplotlib.figure import Figure
# from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plotter
import matplotlib.animation as animation
from matplotlib import style
import mpld3

stopCamera = False
colors = {'neutral':(255, 255, 255), 'angry':(0, 0, 255), 'disgust':(0, 139, 139), 'fear':(125, 125, 125), 'happy':(0, 255, 255), 'sad':(255, 0, 0), 'surprised':(255, 245, 0)}
   
#giving value to each mood
imotions = {0:'angry', 1:'fear', 2:'happy', 3:'sad',
               4:'surprised', 5:'neutral'}

finalArr=[0,0,0,0,0,0]
x_axis=[]
y_axis=[]

ix_axis=[]
iy_axis=[]

pie_arr=[0,0,0,0,0,0]
pie_count=0

lovy=0
p=0
cham=0
champs=0

wwe=1


file = open("example.txt","r+")
file. truncate(0)
file. close()


style.use('fivethirtyeight')

# fig = plt.figure()
# ax1 = fig.add_subplot(1,1,1)

# def animate(i):
#     # global ani
#     graph_data = open('example.txt','r').read()
#     lines = graph_data.split('\n')
#     xs = []
#     ys = []
#     for line in lines:
#         if len(line) > 1:
#             x, y = line.split(',')
#             xs.append(float(x))
#             ys.append(float(y))
#     ax1.clear()
#     xs=xs[-30:]
#     ys=ys[-30:]
#     ax1.set_ylim([0,100])
#     ax1.plot(xs, ys,color='green', linewidth = 2,marker='o', markerfacecolor='blue', markersize=4)
#     plt.xlabel('Time ----->',color='blue')
#     plt.ylabel('Attentiveness',color='blue')
#     plt.tight_layout()

# ani = animation.FuncAnimation(fig, animate, interval=1000)
# plt.tight_layout()
    
    # output = io.BytesIO()
    # FigureCanvasSVG(fig).print_svg(output)
    # return Response(output.getvalue(), mimetype="image/svg+xml")
     
#plt.show()

@app.route('/data', methods=["GET", "POST"])
def data():
    graph_data = open('example.txt','r').read()
    lines = graph_data.split('\n')
    xs = []
    ys = []
    for line in lines:
        if len(line) > 1:
            x, y = line.split(',')
            xs.append(float(x))
            ys.append(float(y))
    
    # xs=xs[-30:]
    ys=ys[-30:]
    
    response=make_response(json.dumps(ys))
    response.content_type='application/json'
    return response
    

workbook = xlsxwriter.Workbook('FinalData.xlsx')
worksheet = workbook.add_worksheet()
worksheet.set_column(1, 1, 20)
worksheet.set_column(0, 1, 20)
worksheet.set_column(2, 1, 20)
workbook.close()


workbook_obj = openpyxl.load_workbook('FinalData.xlsx')
sheet_obj = workbook_obj.active
col1 = 'TIME'
col2 = 'Levels'
col3 = '(Real-time)'
col4 = '(%age)'
col5 = 'Moods'
col6 = '(Patterns)'
col7 = '(Att/In-Att)'
col8 = '(classification)'
sheet_obj.append([col1, col2])
sheet_obj['A1'].font = Font(bold=True)
sheet_obj['B1'].font = Font(bold=True)
sheet_obj.append([col3, col4])
sheet_obj['A2'].font = Font(bold=True)
sheet_obj['B2'].font = Font(bold=True)
sheet_obj['C1']=col5
sheet_obj['D1']=col7
sheet_obj['C1'].font = Font(bold=True)
sheet_obj['D1'].font = Font(bold=True)
sheet_obj['C2']=col6
sheet_obj['D2']=col8
sheet_obj['C2'].font = Font(bold=True)
sheet_obj['D2'].font = Font(bold=True)
workbook_obj.save('FinalData.xlsx')

model = load_model('epoch_75.hdf5')
#Video capturing
face_cascade = cv.CascadeClassifier('haarcascade_frontalface_default.xml')


def convert_dtype(x):
    x_float = x.astype('float32')
    return x_float

def normalize(x):
    x_n = (x - 0)/(255)
    return x_n


def reshape(x):
    x_r = x.reshape((x.shape[0], x.shape[1], x.shape[2], 1))
    return x_r

video_stream = VideoCamera()

@app.route('/show_result/')
def switch_camera():
     video_stream.video.release()     
     return render_template('graphs.html')

figure, axis = plotter.subplots(4, 1,figsize=(15, 15), sharey=True)
plotter.subplots_adjust(hspace=1)


@app.route('/graphs/')
def graphs():
    
    fig = plt.figure()
    plt.plot(x_axis, y_axis,color='green', linewidth = 2,marker='o', markerfacecolor='blue', markersize=3,label='Attentiveness') 
    plt.tight_layout()
    plt.plot(ix_axis, iy_axis,color='red', linewidth = 2,marker='o', markerfacecolor='blue', markersize=3,label='In-Attentiveness') 
    plt.tight_layout()
    plt.xlabel('TIME(in sec) ------> ',color='blue')
    plt.title(' Attentiveness vs In-Attentiveness',color='blue')
    plt.legend(bbox_to_anchor=(1.04,1), loc="upper right")
    plt.tight_layout()
    # output = io.BytesIO()
    # FigureCanvasSVG(fig).print_svg(output)      
    # return Response(output.getvalue(), mimetype="image/svg+xml")

    fig2 = plt.figure(facecolor='yellow')
    #plt.title("Proportionate %age\n" + "of Mood Patterns", bbox={'facecolor':'0.8', 'pad':5})
    ax = fig2.add_axes([0,0,1,1])
    explodeTuple = (0.0, 0.0, 0.0, 0.0, 0.0, 0.06)
    ax.axis('equal')
    langs = ['angry','fear','happy','sad','surprised','neutral']
    students = [x / pie_count for x in pie_arr]
    students = [x *100 for x in students]
    ax.pie(students, explode=explodeTuple,labels = langs,autopct='%1.2f%%',textprops={'color':"black"})
    # output = io.BytesIO()
    # FigureCanvasSVG(fig3).print_svg(output)
    # finalImg = Response(output.getvalue(), mimetype="image/svg+xml")
    
    langs = 'angry','fear','happy','sad','surprised','neutral',
    fig3 = plt.figure()
    fig3.patch.set_facecolor('red')
    plt.rcParams['text.color'] = 'blue'
    my_circle=plt.Circle( (0,0), 0.7, color='black')
    students = [x / pie_count for x in pie_arr]
    students = [x *100 for x in students]
    plt.pie(students, labels=langs)
    p=plt.gcf()
    p.gca().add_artist(my_circle)
    
    fig4 = plt.figure(figsize=(9,4))
    objects = ('angry','fear','happy','sad','surprised','neutral')
    y_pos = np.arange(len(objects))
    performance = pie_arr
    plt.bar(y_pos, performance, align='center', alpha=1,color=['yellow', 'red', 'green', 'blue', 'black','cyan'])
    plt.xticks(y_pos, objects)
    plt.xlabel('Different Moods',color='red')
    plt.ylabel('Mood Levels',color='red')
    plt.title('Overall Mood Patterns',color='red')
    plt.tight_layout()
    # plt.show()

    html1 = mpld3.fig_to_html(fig)
    html2 = mpld3.fig_to_html(fig2)
    html3 = mpld3.fig_to_html(fig3)
    html4 = mpld3.fig_to_html(fig4)
    return (html1 + html2+ html3 + html4)

@app.route('/')
def index():
    return render_template('index.html')

def gen(camera):
    global stopCamera
    global lovy
    global pie_count, pie_arr, finalArr, cham, champs, p, wwe
    
    #Defining colours for each mood
   
    while True:
        img = camera.get_image()
        gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x,y,w,h) in faces:
            roi_gray = gray[y:y+h, x:x+w]
            roi_color = img[y:y+h, x:x+w]
            roi_gray = cv.resize(roi_gray, (48, 48), interpolation = cv.INTER_AREA)
            roi_gray = convert_dtype(np.array([roi_gray]))
            roi_gray = normalize(roi_gray)
            roi_gray = reshape(roi_gray)
            pr = model.predict(roi_gray)[0]
           # print(pr)
            max_emo = np.argmax(pr)
            cv.rectangle(img,(x,y),(x+w,y+h), colors[imotions[max_emo]], 2)
            cv.rectangle(img,(x,y+h),(x+w,y+h+170),(128,128,128), -1)
            cv.rectangle(img,(x,y),(x+w,y+h+130), colors[imotions[max_emo]], 2)
            cv.rectangle(img,(x,y),(x+w,y+h+170), colors[imotions[max_emo]], 2)
            
            counter = 0
            print(lovy)
            lovy=lovy+1
            
            pie_count=pie_count+1
            pie_arr = list(map(add, pr , pie_arr))
            
            finalArr = list(map(add, pr , finalArr))
            
              
                
            
            """
            if lovy==200:
                finalArr = [x / 200 for x in finalArr]
                finalArr = [x * 100 for x in finalArr]
                
                print(finalArr)
                
                if (finalArr[2] >40 or finalArr[5] >40) and (finalArr[0] < 25 and finalArr[3] <25):
                    finalArr=[0,0,0,0,0,0]
                    lovy=0
                    print("ATTTENTIVE")
                    print("\n")
                    root = Tk()
                    root.geometry("150x100") 
                    root.title("ALERT Message")
                    root.eval('tk::PlaceWindow . center')
                    root.configure(background="black")
                    
                    button = Button(root,height = 8, width = 10,bg='red',fg='black', text = 'Attentive',font="comicsansms 15 bold").pack() 
                    start = time() 
                    root.after(2000, root.destroy) 
                    mainloop()  
                    
                else:
                    finalArr=[0,0,0,0,0,0]
                    lovy=0
                    print("INNN-ATTTENTIVE")
                    print("\n")
                    root = Tk()
                    root.geometry("150x100") 
                    root.title("ALERT Message")
                    root.eval('tk::PlaceWindow . center')
                    root.configure(background="black")
    
                    button = Button(root,height = 8, width = 10,bg='red',fg='black', text = 'IN-Attentive',font="comicsansms 15 bold").pack() 
                    start = time() 
                    root.after(2000, root.destroy) 
                    mainloop()  
            """
            
            if lovy%10==0:
                x_axis.append(p)
                ix_axis.append(p)  
                y_axis.append(champs)
                iy_axis.append(cham)
                p=p+1
                
            
            
            for i in range(len(pr)+1):
                
                
                if i!=6:
                    cv.rectangle(img, (x, y+h+counter+7), (x + int(w * pr[i]), y+h+counter+28), colors[imotions[i]], -2)    
                    counter += 20
                    cv.putText(img, str(int(pr[i]*100)), (x + int(w * pr[i]), (y + h +counter+5)), cv.FONT_HERSHEY_SIMPLEX, 0.50,(51,0,0) , 1)
                    if i != 5:
                        cv.putText(img, imotions[i], (x, (y + h +counter+5)), cv.FONT_HERSHEY_SIMPLEX, 0.75,(0,0,0) , 1)
                    else:
                        cv.putText(img, imotions[i], (x, (y + h +counter+5)), cv.FONT_HERSHEY_SIMPLEX, 0.75,(0, 0, 0) , 1)
                
                
                if i==6:
                    counter +=20;
                    
                    if (pr[2] >0.4 or pr[5] >0.4) and (pr[0] < 0.35 and pr[3] <0.35):
            
                        cv.rectangle(img, (x, y+h+counter-4), (x + w, y+h+counter+24),(0,255,0), -2)
                        cv.putText(img,'Attentive', (x+10, (y + h +counter+20)), cv.FONT_HERSHEY_SIMPLEX, 0.75,(0, 0, 0) , 2)
                        if pr[2]>0.5:
                            champs=int(pr[2]*100)
                            col2 = champs
                            col3 ='happy'
                        elif pr[5]>0.5:
                            champs=int(pr[5]*100)
                            col2 = champs
                            col3 ='neutral'
                        else:
                            champs=int(((pr[2]+pr[5])/2)*100)
                            col2 = champs
                            col3 ='neutral'
                        
                        workbook_obj = openpyxl.load_workbook('FinalData.xlsx')
                        sheet_obj = workbook_obj.active
                        col1 = wwe
                        col4 ='Attentive'
                        sheet_obj.append([col1, col2,col3,col4])
                        workbook_obj.save('FinalData.xlsx')
                        
                        file1 = open("example.txt","a")
                        file1. write("\n") 
                        file1.write(str(wwe)+","+str(champs)) 
                        
                        file1.close()
                        
                        wwe=wwe+1
                        
                        cv.putText(img,str(champs)+'%', (x+120, (y + h +counter+20)), cv.FONT_HERSHEY_SIMPLEX, 0.75,(0, 0, 0) , 2)
                        
                    else:
                        cv.rectangle(img, (x, y+h+counter-4), (x + w, y+h+counter+24),(0,0,255), -2)
                        cv.putText(img,'In-Attentive', (x+5, (y + h +counter+20)), cv.FONT_HERSHEY_SIMPLEX, 0.75,(0, 0, 0) , 2)
                        if pr[0]>0.5:
                            cham=int(pr[0]*100)
                            col2 = cham
                            col3 ='Angry'
                        elif pr[3]>0.5:
                            cham=int(pr[3]*100)
                            col2 = cham
                            col3 ='Sad'
                        elif pr[4]>0.5:
                            cham=int(pr[4]*100)
                            col2 = cham
                            col3 ='Surprized'                      
                        else:
                            cham=int(((pr[0]+pr[3]+pr[4])/3)*100)
                            cham=int(pr[3]*100)
                            col2 = cham
                            col3 ='Sad'
                            
                        workbook_obj = openpyxl.load_workbook('FinalData.xlsx')
                        sheet_obj = workbook_obj.active
                        col1 = wwe
                        col4 ='In-Att'
                        sheet_obj.append([col1, col2,col3,col4])
                        workbook_obj.save('FinalData.xlsx')
                        cv.putText(img,str(cham)+'%', (x+150, (y + h +counter+20)), cv.FONT_HERSHEY_SIMPLEX, 0.75,(0, 0, 0) , 2)
                        
                        wwe=wwe+1
            #cv.circle(img, ((x + w//2), (y + h//2)), int(((h*h + w*w)**0.5)//2), colors[imotions[pr]], 2)
            #cv.putText(img, imotions[pr], ((x + w//2), (y + h//2) - int(((h*h + w*w)**0.5)//2)), cv.FONT_HERSHEY_SIMPLEX, 1, colors[imotions[pr]], 1)
        
        #cv.imshow('img',img)
        #print(keypress)
        ret, jpeg = cv.imencode('.jpg', img)
        frame = jpeg.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')

@app.route('/video_feed')
def video_feed():
     return Response(gen(video_stream),
                 mimetype='multipart/x-mixed-replace; boundary=frame')

if __name__ == '__main__':
    app.run(host='127.0.0.1', debug=False, port="5000")
    